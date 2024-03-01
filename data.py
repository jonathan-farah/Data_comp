import nltk
from nltk.chat.util import Chat, reflections
import requests
from bs4 import BeautifulSoup
import openpyxl
def scrape_data_from_website(url):
    # Function to scrape data from a given website
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    
    # Modify this part to extract the specific data you need from the website
    # For example, let's extract the title of the webpage
    title = soup.title.string if soup.title else 'No title found'
    
    return title

def main():
    # Load Excel file
    excel_file_path = 'your_excel_file.xlsx'
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    # Assuming the URLs are in column A starting from row 2
    for row in range(2, sheet.max_row + 1):
        url = sheet.cell(row=row, column=1).value

        if url:
            try:
                # Scrape data from the website
                scraped_data = scrape_data_from_website(url)

                # Assuming you want to store the scraped data in column B
                sheet.cell(row=row, column=2).value = scraped_data

                print(f"Scraped data from {url}: {scraped_data}")
            except Exception as e:
                print(f"Error scraping data from {url}: {e}")

    # Save the updated Excel file
    workbook.save(excel_file_path)

if __name__ == "__main__":
    main()
